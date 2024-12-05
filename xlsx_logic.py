import datetime

from openpyxl import load_workbook
import config


def find_name(name):
    wb = load_workbook(filename=config.path_to_main)
    ws = wb['main']
    found = []

    for row in ws.iter_rows(max_row=1500, min_col=8, max_col=8):
        for cell in row:
            if cell.value:
                if name in cell.value:
                    data = {'name': ws.cell(row=cell.row, column=8).value}

                    localization = ws.cell(row=cell.row, column=2).value
                    if localization:
                        data['localization'] = localization
                    else:
                        data['localization'] = 'нет данных'

                    company = ws.cell(row=cell.row, column=5).value
                    if company:
                        data['company'] = company
                    else:
                        data['company'] = 'нет данных'

                    phase = ws.cell(row=cell.row, column=10).value
                    if phase:
                        data['phase'] = phase
                    else:
                        data['phase'] = 'нет данных'

                    group = ws.cell(row=cell.row, column=7).value
                    if group:
                        data['group'] = group
                    else:
                        data['group'] = 'нет данных'

                    found.append(data)
            else:
                continue
    return found


def find_name_n_ready(name):
    wb = load_workbook(filename=config.path_to_main)
    ws = wb['n_ready']
    found = []
    for row in ws.iter_rows(max_row=500, min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                if name in cell.value:
                    data = {'name': ws.cell(row=cell.row, column=1).value}

                    localization = ws.cell(row=cell.row, column=2).value
                    if localization:
                        data['localization'] = localization
                    else:
                        data['localization'] = 'нет данных'

                    company = ws.cell(row=cell.row, column=3).value
                    if company:
                        data['company'] = company
                    else:
                        data['company'] = 'нет данных'

                    reason = ws.cell(row=cell.row, column=4).value
                    if reason:
                        data['reason'] = reason
                    else:
                        data['reason'] = 'нет данных'

                    again = ws.cell(row=cell.row, column=5).value
                    if again:
                        data['again'] = again
                    else:
                        data['again'] = 'нет данных'

                    found.append(data)
            else:
                continue
    return found


def find_tel_name(name):
    wb = load_workbook(filename=config.path_to_base)
    ws = wb['Общий список']
    found = []

    for row in ws.iter_rows(max_row=3000, min_col=2, max_col=2):
        for cell in row:
            if cell.value:
                if name in cell.value:
                    data = {'name': ws.cell(row=cell.row, column=2).value}
                    bd = ws.cell(row=cell.row, column=4).value
                    if bd and isinstance(bd, datetime.datetime):
                        data['bd'] = bd.strftime("%d.%m.%Y")
                    elif bd:
                        data['bd'] = bd
                    else:
                        data['bd'] = 'нет данных'

                    localization = ws.cell(row=cell.row, column=12).value
                    if localization:
                        data['localization'] = localization
                    else:
                        data['localization'] = 'нет данных'

                    tel = ws.cell(row=cell.row, column=27).value
                    if tel:
                        data['tel'] = tel
                    else:
                        data['tel'] = 'нет данных'

                    clinic = ws.cell(row=cell.row, column=8).value
                    if clinic:
                        data['clinic'] = clinic
                    else:
                        data['clinic'] = 'нет данных'

                    found.append(data)
                else:
                    continue

    return found

def find_tel_cli(name):
    wb = load_workbook(filename=config.path_to_contact)
    ws = wb['main']
    found = []

    for row in ws.iter_rows(max_row=30, min_col=2, max_col=2):
        for cell in row:
            if cell.value:
                if name in cell.value:
                    data = {'cli': ws.cell(row=cell.row, column=2).value}
                    data['adr'] = ws.cell(row=cell.row, column=3).value
                    data['con'] = ws.cell(row=cell.row, column=4).value
                    data['name'] = ws.cell(row=cell.row, column=1).value



                    found.append(data)
                else:
                    continue

    return found

def find_name_reamp(name):
    wb = load_workbook(filename=config.path_to_reamp)
    ws = wb['Лист1']
    found = []

    for row in ws.iter_rows(max_row=1500, min_col=7, max_col=7):
        for cell in row:
            if cell.value:
                if name in cell.value:
                    data = {'name': ws.cell(row=cell.row, column=7).value}

                    localization = ws.cell(row=cell.row, column=2).value
                    if localization:
                        data['localization'] = localization
                    else:
                        data['localization'] = 'нет данных'

                    company = ws.cell(row=cell.row, column=4).value
                    if company:
                        data['company'] = company
                    else:
                        data['company'] = 'нет данных'

                    phase = ws.cell(row=cell.row, column=9).value
                    if phase:
                        data['phase'] = phase
                    else:
                        data['phase'] = 'нет данных'

                    group = ws.cell(row=cell.row, column=6).value
                    if group:
                        data['group'] = group
                    else:
                        data['group'] = 'нет данных'

                    found.append(data)
            else:
                continue
    return found



if __name__ == "__main__":
    # print(find_name('Генинсон'))
    # print(find_tel_name('Генинсон'))
    # print(find_name_n_ready('Махмудов Муса'))
    print(find_tel_cli("Л"))
